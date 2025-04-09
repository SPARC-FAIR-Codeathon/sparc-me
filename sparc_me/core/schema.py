import json
import pandas as pd
from pathlib import Path
from typing import Dict
import openpyxl
import os
import math

from sparc_me.core.utils import CaseInsensitiveDict, validate_metadata_file

from xlrd import XLRDError

from jsonschema import validate
from jsonschema.exceptions import ValidationError
# from jsonschema import Draft202012Validator
from jsonschema import Draft7Validator
from jsonschema.exceptions import best_match

current_dir = Path(__file__).parent.resolve()
resources_dir = Path.joinpath(current_dir, "../resources")


def convert_version_format(version):
    """
    Convert version format
    :param version: dataset/template version
    :type version: string
    :return: version in the converted format
    :rtype:
    """
    version = version.replace(".", "_")

    if "_" not in version:
        version = version + "_0_0"

    return version


class Validator(object):
    def __init__(self):
        self._version = None
        self._metadata_file = None
        self._schema_ref = None

    def validate_dataset(self, dataset):

        metadata_files = dataset.list_metadata_files(version="2.0.0", print_list=False)
        currently_support_metadata_files = ['dataset_description', 'manifest', 'samples', 'subjects']
        schema = Schema()

        for metadata_file in currently_support_metadata_files:
            metadata = dataset.get_metadata(metadata_file)
            data = schema.load_data(metadata.metadata_file_path)
            self.validate(data, metadata_file=metadata_file, version=metadata.version)

    def validate(self, data, metadata_file, version):
        """
        Validate data instance
        :param data: Target data
        :type data: dict or list
        :param metadata_file: metadata filename
        :type metadata_file: string
        :param version: dataset version
        :type version: string
        :return:
        :rtype:
        """
        self._metadata_file = metadata_file
        self._version = convert_version_format(version)
        self._load_reference_schema()
        if isinstance(data, dict):
            self._execute(data)
        elif isinstance(data, list):
            for idx in range(len(data)):
                self._execute(data[idx])
        else:
            print("Input data type invalid")

    def _load_reference_schema(self):
        """
        Load the reference schema which will be used to validate the target metadata
        :return:
        :rtype:
        """
        version = "version_" + self._version
        filename = self._metadata_file + ".json"
        schema_path = resources_dir / "templates" / version / "schema" / filename
        if not schema_path.exists():
            raise ValueError("Reference schema not found!")

        with open(schema_path) as file:
            schema_details = json.load(file)

        schema = {"type": schema_details.get("type"), "properties": {}, "required": schema_details.get("required")}
        for element, value in schema_details["properties"].items():
            data_type = value.get("type")
            schema["properties"][element] = {
                "type": data_type
            }

        self._schema_ref = schema

    def _execute(self, data):
        """
        Run the validation
        :param data: target data
        :type data: dict
        :return: validation status
        :rtype: bool
        """
        print("Target instance: " + str(data))
        # validate(instance=instance, schema=self._schema_ref)
        try:
            validate(instance=data, schema=self._schema_ref)
            print("Validation: Passed")
            return True
        except ValidationError:
            print(best_match(Draft7Validator(self._schema_ref).iter_errors(data)).message)
            print("Validation: Failed")
            return False
        except Exception as e:
            print(str(e))
            print("Validation: Failed")
            return False


class Schema(object):
    def __init__(self):
        self._version = None
        self._schema = dict()
        self._schemas = dict()
        self._metadata_files = list()

        self._schema_dir = Path()

        self._column_based = ["dataset_description", "code_description"]

    @staticmethod
    def get_default_schema(version, metadata_file):
        version = convert_version_format(version)
        version = "version_" + version
        filename = metadata_file + ".json"
        schema_path = resources_dir / "templates" / version / "schema" / filename
        if not schema_path.exists():
            raise ValueError("Reference schema not found!")

        with open(schema_path) as file:
            schema = json.load(file)

        return schema

    def get_schema(self, metadata_file, version="2.0.0", print_schema=True, required_only=True, name_only=True):
        """
        get a schema via metadata_file/metadate file name

        :param metadata_file: the metadata file name
        :type metadata_file: str
        :param version: "2.0.0"|"1.2.3"
        :type version: str
        :return: dict
        """
        metadata_file = validate_metadata_file(metadata_file, version)
        filename = metadata_file + ".json"
        if version == "2.0.0":
            schema_path = resources_dir / "templates" / "version_2_0_0" / "schema" / filename
        elif version == "1.2.3":
            schema_path = resources_dir / "templates" / "version_1_2_3" / "schema" / filename
        else:
            msg = "Please provide a correct version 2.0.0 or 1.2.3"
            raise ValueError(msg)

        with open(schema_path, 'r') as file:
            schema_json: Dict = json.load(file)

            if required_only:
                if name_only:
                    if print_schema:
                        print(f"The required elements for {metadata_file} name_only:")
                        print(json.dumps(schema_json.get('required'), indent=4))
                    return schema_json.get('required')
                else:
                    required_items = []
                    for key, value in schema_json.get('properties').items():
                        if "required" in value and value["required"] == "Y":
                            required_items.append({key: value})
                    if print_schema:
                        print(f"The required elements for {metadata_file}:")
                        print(json.dumps(required_items, indent=4))
                    return required_items
            else:
                print(json.dumps(schema_json.get('properties'), indent=4))
            return CaseInsensitiveDict(schema_json.get('properties'))

    def get(self):
        return self._schema

    def set_schema(self, schema):
        self._schema = schema

    def add_property(self, property_name, property):
        self._schema["properties"][property_name] = property
        required = property.get("required")
        if required == 'Y':
            self._schema["required"].append(property_name)

    def load_data(self, path):
        """
        Load in a metadata
        :param path: path to the metadata file
        :type path: string
        :return: data instance
        :rtype: dict or list
        """
        path = Path(path)
        filename = path.stem
        try:
            data_pd = pd.read_excel(path)
        except XLRDError:
            data_pd = pd.read_excel(path, engine='openpyxl')

        data_dict = data_pd.to_dict()
        return self.generate_validate_data(filename, data_dict)

    def generate_validate_data(self, filename, data_dict):
        if filename in self._column_based:
            data = {}
            elements = data_dict.get("Metadata element")
            values = data_dict.get("Value")
            length = len(elements)
            for idx in range(length):
                element = elements[idx]
                value = values[idx]
                try:
                    if math.isnan(value):
                        continue
                except:
                    pass

                data[element] = value
        else:
            data = list()

            elements = list(data_dict)
            length = len(data_dict[elements[0]])
            for idx in range(length):
                instance = {}
                for element in elements:
                    value = data_dict[element][idx]

                    try:
                        if math.isnan(value):
                            continue
                    except:
                        pass

                    instance[element] = value

                if instance:
                    data.append(instance)

        return data

    def generate_from_template(self, file, save_dir):
        """
        Generate schema from template
        (the element_description.xlsx file in the template folder)
        :param file: path to the template
        :type file: string
        :param save_dir: directory to save the schema files
        :type save_dir: string
        :return:
        :rtype:
        """
        wb = openpyxl.load_workbook(file)
        sheetnames = wb.sheetnames

        for sheet in sheetnames:
            try:
                schema_pd = pd.read_excel(file, sheet_name=sheet)
            except XLRDError:
                schema_pd = pd.read_excel(file, sheet_name=sheet, engine='openpyxl')

            schema = {"type": "object", "properties": {}, "required": []}
            required_list = list()

            for index, row in schema_pd.iterrows():
                element = row.get("Element")
                required = row.get("Required")
                data_type = row.get("Type")
                description = row.get("Description")
                example = row.get("Example")

                if required == 'Y':
                    required_list.append(element)

                # data_type = type(example).__name__

                try:
                    if math.isnan(example):
                        example = None
                    if math.isnan(description):
                        description = None
                except:
                    pass

                schema["properties"][element] = {
                    "type": data_type,
                    "required": required,
                    "description": description,
                    "example": example
                }

            schema["required"] = required_list
            self._schemas[sheet] = schema

            self.save(save_dir, schema, sheet)

    def convert_schema_excel_to_json(self, source_path, dest_path):
        """
        :param source_path: The Excel schema file path
        :type source_path: str
        :param dest_path: The converted Json file store path
        :type dest_path: str
        """
        wb = openpyxl.load_workbook(source_path)
        sheets = wb.sheetnames

        schema = dict()
        for sheet in sheets:
            schema[sheet] = dict()
            try:
                element_description = pd.read_excel(source_path, sheet_name=sheet)
            except XLRDError:
                element_description = pd.read_excel(source_path, sheet_name=sheet, engine='openpyxl')

            element_description = element_description.where(pd.notnull(element_description), None)

            for index, row in element_description.iterrows():
                element = row["Element"]
                schema[sheet][element] = dict()
                schema[sheet][element]["Required"] = row["Required"]
                schema[sheet][element]["Type"] = row["Type"]
                schema[sheet][element]["Description"] = row["Description"]
                schema[sheet][element]["Example"] = row["Example"]

        with open(dest_path, 'w') as f:
            json.dump(schema, f, indent=4)

    def save(self, save_dir, schema, metadata_file):
        """
        Save schema
        :param save_dir: path to the destination directory
        :type save_dir: string
        :param schema: metadata schema
        :type schema: dict
        :param metadata_file: metadata metadata_file (filename)
        :type metadata_file: string
        :return:
        :rtype:
        """
        save_dir = Path(save_dir)
        if not save_dir.exists():
            os.makedirs(save_dir)

        filename = '{sheet}.json'.format(sheet=metadata_file)

        save_path = Path.joinpath(save_dir, filename)
        with open(save_path, 'w', encoding='utf-8') as f:
            json.dump(schema, f, indent=4)


if __name__ == '__main__':
    # schema_xlsm = Path.joinpath(current_dir, "../resources/templates/version_1_2_3/schema.xlsx")
    # save_dir = Path.joinpath(current_dir, "../resources/templates/version_1_2_3/schema")
    schema_xlsm = Path.joinpath(current_dir, "../resources/templates/version_2_0_0/schema.xlsx")
    save_dir = Path.joinpath(current_dir, "../resources/templates/version_2_0_0/schema")

    schema = Schema()
    schema.generate_from_template(schema_xlsm, save_dir=save_dir)
    schema.convert_schema_excel_to_json(schema_xlsm, "../resources/templates/version_2_0_0/schema.json")

    # instance = {"Name": "test", "Description": "test", "Keywords":"sassf"}
    # validator = Validator()
    # validator.validate(instance, metadata_file="dataset_description", version="1.2.3")
